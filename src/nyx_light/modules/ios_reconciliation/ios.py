"""
Modul A9 — IOS Usklađivanja (Izvod Otvorenih Stavki)

Kompletna funkcionalnost:
  1. Generiranje IOS obrazaca (Excel i PDF format)
  2. Praćenje povrata putem email-a
  3. Automatsko mapiranje razlika u Excel radnu listu
  4. Prijedlog kompenzacijskih knjiženja

IOS proces:
  Klijent → Generira IOS za svakog partnera →
  Šalje emailom → Partner potvrdi/ospori →
  Mapira razlike → Predlaže knjiženja
"""

import io
import json
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("nyx_light.modules.ios")


class IOSStavka:
    """Jedna otvorena stavka u IOS obrascu."""

    def __init__(
        self,
        broj_dokumenta: str,
        datum_dokumenta: str,
        datum_dospijeca: str,
        opis: str,
        duguje: float = 0.0,
        potrazuje: float = 0.0,
    ):
        self.broj_dokumenta = broj_dokumenta
        self.datum_dokumenta = datum_dokumenta
        self.datum_dospijeca = datum_dospijeca
        self.opis = opis
        self.duguje = duguje
        self.potrazuje = potrazuje

    @property
    def saldo(self) -> float:
        return round(self.duguje - self.potrazuje, 2)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "broj_dokumenta": self.broj_dokumenta,
            "datum_dokumenta": self.datum_dokumenta,
            "datum_dospijeca": self.datum_dospijeca,
            "opis": self.opis,
            "duguje": self.duguje,
            "potrazuje": self.potrazuje,
            "saldo": self.saldo,
        }


class IOSObrazac:
    """IOS obrazac za jednog partnera."""

    def __init__(
        self,
        client_id: str,
        client_name: str,
        client_oib: str,
        partner_name: str,
        partner_oib: str,
        datum_od: str,
        datum_do: str,
    ):
        self.client_id = client_id
        self.client_name = client_name
        self.client_oib = client_oib
        self.partner_name = partner_name
        self.partner_oib = partner_oib
        self.datum_od = datum_od
        self.datum_do = datum_do
        self.stavke: List[IOSStavka] = []
        self.status = "draft"
        self.created_at = datetime.now().isoformat()
        self.sent_at: Optional[str] = None
        self.response_at: Optional[str] = None
        self.partner_stavke: List[IOSStavka] = []
        self.razlike: List[Dict[str, Any]] = []

    def add_stavka(self, stavka: IOSStavka):
        self.stavke.append(stavka)

    @property
    def ukupno_duguje(self) -> float:
        return round(sum(s.duguje for s in self.stavke), 2)

    @property
    def ukupno_potrazuje(self) -> float:
        return round(sum(s.potrazuje for s in self.stavke), 2)

    @property
    def saldo(self) -> float:
        return round(self.ukupno_duguje - self.ukupno_potrazuje, 2)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "client_id": self.client_id,
            "client_name": self.client_name,
            "client_oib": self.client_oib,
            "partner_name": self.partner_name,
            "partner_oib": self.partner_oib,
            "period": f"{self.datum_od} — {self.datum_do}",
            "datum_od": self.datum_od,
            "datum_do": self.datum_do,
            "stavke": [s.to_dict() for s in self.stavke],
            "ukupno_duguje": self.ukupno_duguje,
            "ukupno_potrazuje": self.ukupno_potrazuje,
            "saldo": self.saldo,
            "status": self.status,
            "created_at": self.created_at,
            "sent_at": self.sent_at,
            "response_at": self.response_at,
            "razlike_count": len(self.razlike),
        }


class IOSReconciliation:
    """
    IOS usklađivanja — generiranje obrazaca, tracking, razlike.

    Metode:
      - generate_ios_form() → IOSObrazac s Excel exportom
      - generate_excel() → bytes (Excel datoteka)
      - track_responses() → status praćenja
      - map_differences() → lista razlika s prijedlogom knjiženja
      - generate_difference_report() → Excel s razlikama
    """

    def __init__(self, export_dir: str = "data/exports/ios"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        self._forms: Dict[str, IOSObrazac] = {}
        self._tracking: Dict[str, Dict[str, Any]] = {}

    def generate_ios_form(
        self,
        client_id: str,
        client_name: str = "",
        client_oib: str = "",
        partner_name: str = "",
        partner_oib: str = "",
        datum_od: str = "",
        datum_do: str = "",
        stavke: Optional[List[Dict]] = None,
    ) -> Dict[str, Any]:
        """Generiraj IOS obrazac za klijenta i partnera."""
        if not datum_od:
            datum_od = f"{datetime.now().year}-01-01"
        if not datum_do:
            datum_do = datetime.now().strftime("%Y-%m-%d")

        form = IOSObrazac(
            client_id=client_id,
            client_name=client_name or f"Klijent {client_id}",
            client_oib=client_oib,
            partner_name=partner_name or f"Partner {partner_oib}",
            partner_oib=partner_oib,
            datum_od=datum_od,
            datum_do=datum_do,
        )

        # Add stavke
        if stavke:
            for s in stavke:
                form.add_stavka(IOSStavka(
                    broj_dokumenta=s.get("broj_dokumenta", ""),
                    datum_dokumenta=s.get("datum_dokumenta", ""),
                    datum_dospijeca=s.get("datum_dospijeca", ""),
                    opis=s.get("opis", ""),
                    duguje=float(s.get("duguje", 0)),
                    potrazuje=float(s.get("potrazuje", 0)),
                ))

        form_id = f"IOS-{client_id}-{partner_oib}-{datum_do}"
        self._forms[form_id] = form

        result = form.to_dict()
        result["form_id"] = form_id
        result["format"] = "xlsx"

        # Generate Excel
        excel_path = self._generate_excel(form, form_id)
        if excel_path:
            result["excel_path"] = str(excel_path)

        logger.info("IOS generiran: %s (%d stavki, saldo: %.2f EUR)",
                    form_id, len(form.stavke), form.saldo)
        return result

    def _generate_excel(self, form: IOSObrazac, form_id: str) -> Optional[Path]:
        """Generiraj Excel IOS obrazac."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            logger.warning("openpyxl nije instaliran — Excel export nedostupan")
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IOS"

        # Styles
        bold = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center")
        thin = Side(border_style="thin")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "IZVOD OTVORENIH STAVKI (IOS)"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center

        # Client info
        ws["A3"] = "Tvrtka:"
        ws["B3"] = form.client_name
        ws["A4"] = "OIB:"
        ws["B4"] = form.client_oib
        ws["A5"] = "Partner:"
        ws["B5"] = form.partner_name
        ws["A6"] = "Partner OIB:"
        ws["B6"] = form.partner_oib
        ws["A7"] = "Razdoblje:"
        ws["B7"] = f"{form.datum_od} — {form.datum_do}"
        ws["A8"] = "Datum izrade:"
        ws["B8"] = datetime.now().strftime("%d.%m.%Y.")

        for r in range(3, 9):
            ws[f"A{r}"].font = bold

        # Headers
        headers = ["R.br.", "Broj dokumenta", "Datum dokumenta",
                   "Datum dospijeća", "Opis", "Duguje (EUR)", "Potražuje (EUR)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # Data
        for i, stavka in enumerate(form.stavke, 1):
            row = 10 + i
            ws.cell(row=row, column=1, value=i).border = border
            ws.cell(row=row, column=2, value=stavka.broj_dokumenta).border = border
            ws.cell(row=row, column=3, value=stavka.datum_dokumenta).border = border
            ws.cell(row=row, column=4, value=stavka.datum_dospijeca).border = border
            ws.cell(row=row, column=5, value=stavka.opis).border = border
            c_dug = ws.cell(row=row, column=6, value=stavka.duguje)
            c_dug.number_format = "#,##0.00"
            c_dug.border = border
            c_pot = ws.cell(row=row, column=7, value=stavka.potrazuje)
            c_pot.number_format = "#,##0.00"
            c_pot.border = border

        # Totals
        total_row = 11 + len(form.stavke)
        ws.merge_cells(f"A{total_row}:E{total_row}")
        ws[f"A{total_row}"] = "UKUPNO:"
        ws[f"A{total_row}"].font = bold
        ws[f"A{total_row}"].alignment = Alignment(horizontal="right")
        c = ws.cell(row=total_row, column=6, value=form.ukupno_duguje)
        c.font = bold
        c.number_format = "#,##0.00"
        c.border = border
        c = ws.cell(row=total_row, column=7, value=form.ukupno_potrazuje)
        c.font = bold
        c.number_format = "#,##0.00"
        c.border = border

        saldo_row = total_row + 1
        ws.merge_cells(f"A{saldo_row}:E{saldo_row}")
        ws[f"A{saldo_row}"] = "SALDO:"
        ws[f"A{saldo_row}"].font = Font(bold=True, size=12)
        ws[f"A{saldo_row}"].alignment = Alignment(horizontal="right")
        c = ws.cell(row=saldo_row, column=6, value=form.saldo)
        c.font = Font(bold=True, size=12, color="FF0000" if form.saldo != 0 else "000000")
        c.number_format = "#,##0.00"

        # Confirmation section
        conf_row = saldo_row + 3
        ws[f"A{conf_row}"] = "POTVRDA USKLAĐENJA:"
        ws[f"A{conf_row}"].font = bold
        ws[f"A{conf_row + 1}"] = "□ Potvrđujem gornji saldo"
        ws[f"A{conf_row + 2}"] = "□ Osporavam — priložen korigirani IOS"
        ws[f"A{conf_row + 4}"] = "Potpis i pečat: ___________________"
        ws[f"A{conf_row + 5}"] = "Datum: ___________________"

        # Column widths
        widths = [6, 18, 16, 16, 30, 15, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # Save
        filepath = self.export_dir / f"{form_id}.xlsx"
        wb.save(filepath)
        return filepath

    def track_responses(self, form_id: str) -> Dict[str, Any]:
        """Prati status odgovora na IOS."""
        form = self._forms.get(form_id)
        if not form:
            return {"form_id": form_id, "status": "not_found"}

        tracking = self._tracking.get(form_id, {
            "sent": form.sent_at is not None,
            "response_received": form.response_at is not None,
            "days_waiting": 0,
            "reminders_sent": 0,
        })

        if form.sent_at and not form.response_at:
            sent_date = datetime.fromisoformat(form.sent_at)
            tracking["days_waiting"] = (datetime.now() - sent_date).days
            if tracking["days_waiting"] > 15:
                tracking["action_needed"] = "Poslati reminder (15+ dana)"
            elif tracking["days_waiting"] > 30:
                tracking["action_needed"] = "HITNO — 30+ dana bez odgovora"

        return {"form_id": form_id, **tracking}

    def mark_sent(self, form_id: str, email_to: str = "") -> Dict[str, Any]:
        """Označi IOS kao poslan."""
        form = self._forms.get(form_id)
        if not form:
            return {"error": f"IOS {form_id} ne postoji"}
        form.sent_at = datetime.now().isoformat()
        form.status = "sent"
        self._tracking[form_id] = {
            "sent": True, "sent_to": email_to,
            "response_received": False, "reminders_sent": 0,
        }
        return {"form_id": form_id, "status": "sent", "sent_at": form.sent_at}

    def receive_response(
        self,
        form_id: str,
        confirmed: bool,
        partner_stavke: Optional[List[Dict]] = None,
    ) -> Dict[str, Any]:
        """Zaprimi odgovor partnera na IOS."""
        form = self._forms.get(form_id)
        if not form:
            return {"error": f"IOS {form_id} ne postoji"}

        form.response_at = datetime.now().isoformat()

        if confirmed:
            form.status = "confirmed"
            return {"form_id": form_id, "status": "confirmed", "razlike": []}

        # Partner osporio — mapiranje razlika
        form.status = "disputed"
        if partner_stavke:
            for s in partner_stavke:
                form.partner_stavke.append(IOSStavka(
                    broj_dokumenta=s.get("broj_dokumenta", ""),
                    datum_dokumenta=s.get("datum_dokumenta", ""),
                    datum_dospijeca=s.get("datum_dospijeca", ""),
                    opis=s.get("opis", ""),
                    duguje=float(s.get("duguje", 0)),
                    potrazuje=float(s.get("potrazuje", 0)),
                ))

        razlike = self.map_differences(form_id)
        return {"form_id": form_id, "status": "disputed", "razlike": razlike}

    def map_differences(self, form_id: str) -> List[Dict[str, Any]]:
        """Mapiraj razlike između naših i partnerovih stavki."""
        form = self._forms.get(form_id)
        if not form:
            return []

        our_docs = {s.broj_dokumenta: s for s in form.stavke}
        partner_docs = {s.broj_dokumenta: s for s in form.partner_stavke}

        razlike = []

        # Documents we have but partner doesn't
        for doc_id, stavka in our_docs.items():
            if doc_id not in partner_docs:
                razlike.append({
                    "tip": "nedostaje_kod_partnera",
                    "broj_dokumenta": doc_id,
                    "nas_iznos": stavka.saldo,
                    "partner_iznos": 0.0,
                    "razlika": stavka.saldo,
                    "opis": f"Dokument {doc_id} postoji kod nas ali ne kod partnera",
                    "prijedlog": "Provjeriti je li dokument dostavljen partneru",
                })

        # Documents partner has but we don't
        for doc_id, stavka in partner_docs.items():
            if doc_id not in our_docs:
                razlike.append({
                    "tip": "nedostaje_kod_nas",
                    "broj_dokumenta": doc_id,
                    "nas_iznos": 0.0,
                    "partner_iznos": stavka.saldo,
                    "razlika": -stavka.saldo,
                    "opis": f"Dokument {doc_id} postoji kod partnera ali ne kod nas",
                    "prijedlog": "Zatražiti kopiju dokumenta od partnera",
                })

        # Amount differences
        for doc_id in our_docs:
            if doc_id in partner_docs:
                our = our_docs[doc_id]
                their = partner_docs[doc_id]
                diff = round(our.saldo - their.saldo, 2)
                if abs(diff) > 0.01:
                    razlike.append({
                        "tip": "razlika_u_iznosu",
                        "broj_dokumenta": doc_id,
                        "nas_iznos": our.saldo,
                        "partner_iznos": their.saldo,
                        "razlika": diff,
                        "opis": f"Razlika u iznosu za {doc_id}: {diff:.2f} EUR",
                        "prijedlog": "Provjeriti izvorne dokumente" if abs(diff) > 100
                        else "Moguća tečajna razlika ili zaokruživanje",
                    })

        form.razlike = razlike
        return razlike

    def generate_difference_report(self, form_id: str) -> Optional[Path]:
        """Generiraj Excel izvještaj o razlikama."""
        form = self._forms.get(form_id)
        if not form or not form.razlike:
            return None

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        except ImportError:
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Razlike"

        bold = Font(bold=True)
        red_fill = PatternFill(start_color="FFCCCC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFCC", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(border_style="thin")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)

        ws.merge_cells("A1:F1")
        ws["A1"] = f"IOS RAZLIKE — {form.partner_name} ({form.partner_oib})"
        ws["A1"].font = Font(bold=True, size=13)

        ws["A2"] = f"Razdoblje: {form.datum_od} — {form.datum_do}"
        ws["A3"] = f"Ukupna razlika: {sum(r['razlika'] for r in form.razlike):.2f} EUR"
        ws["A3"].font = Font(bold=True, color="FF0000")

        headers = ["Tip", "Broj dokumenta", "Naš iznos", "Partner iznos", "Razlika", "Prijedlog"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for i, r in enumerate(form.razlike, 1):
            row = 5 + i
            tip_labels = {
                "nedostaje_kod_partnera": "⚠ Nedostaje kod partnera",
                "nedostaje_kod_nas": "⚠ Nedostaje kod nas",
                "razlika_u_iznosu": "≠ Razlika u iznosu",
            }
            ws.cell(row=row, column=1, value=tip_labels.get(r["tip"], r["tip"])).border = border
            ws.cell(row=row, column=2, value=r["broj_dokumenta"]).border = border
            c = ws.cell(row=row, column=3, value=r["nas_iznos"])
            c.number_format = "#,##0.00"
            c.border = border
            c = ws.cell(row=row, column=4, value=r["partner_iznos"])
            c.number_format = "#,##0.00"
            c.border = border
            c = ws.cell(row=row, column=5, value=r["razlika"])
            c.number_format = "#,##0.00"
            c.border = border
            c.fill = red_fill if abs(r["razlika"]) > 100 else yellow_fill
            ws.cell(row=row, column=6, value=r["prijedlog"]).border = border

        widths = [25, 18, 14, 14, 14, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        filepath = self.export_dir / f"{form_id}_razlike.xlsx"
        wb.save(filepath)
        return filepath

    def get_all_forms(self, client_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """Dohvati sve IOS obrasce."""
        forms = []
        for fid, form in self._forms.items():
            if client_id and form.client_id != client_id:
                continue
            d = form.to_dict()
            d["form_id"] = fid
            forms.append(d)
        return forms

    def get_stats(self) -> Dict[str, Any]:
        statuses = {}
        for f in self._forms.values():
            statuses[f.status] = statuses.get(f.status, 0) + 1
        return {
            "total_forms": len(self._forms),
            "by_status": statuses,
            "total_razlike": sum(len(f.razlike) for f in self._forms.values()),
        }
