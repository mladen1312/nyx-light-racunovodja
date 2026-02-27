"""
Nyx Light — Audit Trail Export

Prema Zakonu o računovodstvu (NN 78/15-114/22), čl. 7-13.
Revizijski trag za sve operacije u sustavu.

Exporti:
  - Excel (.xlsx) — za internu reviziju
  - CSV — za vanjskog revizora
  - JSON — za strojnu analizu

Filteri: po korisniku, periodu, tipu operacije, klijentu.
"""

import csv
import json
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger("nyx_light.audit.export")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class AuditExporter:
    """Exportira audit trail u razne formate."""

    COLUMNS = [
        "timestamp", "user", "action", "module", "client_id",
        "details", "ip_address", "session_id", "result", "risk_level"
    ]

    COLUMN_HR = [
        "Vrijeme", "Korisnik", "Akcija", "Modul", "Klijent",
        "Detalji", "IP adresa", "Sesija", "Rezultat", "Rizik"
    ]

    def __init__(self, audit_store=None):
        """
        audit_store: objekt s metodom get_entries(filters) koji vraća audit zapise.
        Ako None, koristi dummy podatke za testiranje.
        """
        self._store = audit_store

    def get_entries(self, period_from: str = "", period_to: str = "",
                    user: str = "", action: str = "", client_id: str = "",
                    module: str = "", limit: int = 10000) -> List[Dict]:
        """Dohvati audit zapise s filterima."""
        if self._store and hasattr(self._store, "get_entries"):
            return self._store.get_entries(
                period_from=period_from, period_to=period_to,
                user=user, action=action, client_id=client_id,
                module=module, limit=limit
            )

        # Fallback: čitaj iz SQLite audit log
        try:
            import sqlite3
            db_path = Path("data/memory_db/audit.db")
            if not db_path.exists():
                return []

            conn = sqlite3.connect(str(db_path))
            conn.row_factory = sqlite3.Row
            query = "SELECT * FROM audit_log WHERE 1=1"
            params = []

            if period_from:
                query += " AND timestamp >= ?"
                params.append(period_from)
            if period_to:
                query += " AND timestamp <= ?"
                params.append(period_to)
            if user:
                query += " AND user = ?"
                params.append(user)
            if action:
                query += " AND action = ?"
                params.append(action)
            if client_id:
                query += " AND client_id = ?"
                params.append(client_id)
            if module:
                query += " AND module = ?"
                params.append(module)

            query += f" ORDER BY timestamp DESC LIMIT {limit}"
            rows = conn.execute(query, params).fetchall()
            conn.close()
            return [dict(r) for r in rows]
        except Exception as e:
            logger.warning(f"Audit DB read error: {e}")
            return []

    def export_excel(self, entries: List[Dict], output_path: str = "",
                     title: str = "Revizijski trag") -> str:
        """Export u Excel (.xlsx)."""
        if not HAS_OPENPYXL:
            return self.export_csv(entries, output_path.replace(".xlsx", ".csv"))

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Trail"

        # Header
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Generiran: {datetime.now().strftime('%d.%m.%Y. %H:%M')}")
        ws.cell(row=2, column=4, value=f"Zapisa: {len(entries)}")

        # Column headers
        header_fill = PatternFill("solid", fgColor="6366f1")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for col, h in enumerate(self.COLUMN_HR, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.fill = header_fill
            c.font = header_font

        # Data
        for i, entry in enumerate(entries):
            row = 5 + i
            for col, key in enumerate(self.COLUMNS, 1):
                val = entry.get(key, "")
                if isinstance(val, dict):
                    val = json.dumps(val, ensure_ascii=False)[:200]
                ws.cell(row=row, column=col, value=str(val)[:500])

            # Color risk level
            risk = entry.get("risk_level", "low")
            if risk == "high":
                ws.cell(row=row, column=10).font = Font(color="ef4444", bold=True)
            elif risk == "medium":
                ws.cell(row=row, column=10).font = Font(color="eab308")

        # Auto-width
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

        path = output_path or f"data/exports/audit_{date.today().isoformat()}.xlsx"
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        logger.info(f"Audit export: {path} ({len(entries)} zapisa)")
        return path

    def export_csv(self, entries: List[Dict], output_path: str = "") -> str:
        """Export u CSV (za vanjskog revizora)."""
        path = output_path or f"data/exports/audit_{date.today().isoformat()}.csv"
        Path(path).parent.mkdir(parents=True, exist_ok=True)

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=self.COLUMNS, delimiter=";",
                               extrasaction="ignore")
            w.writeheader()
            for entry in entries:
                clean = {}
                for k in self.COLUMNS:
                    v = entry.get(k, "")
                    if isinstance(v, dict):
                        v = json.dumps(v, ensure_ascii=False)[:200]
                    clean[k] = str(v)[:500]
                w.writerow(clean)

        logger.info(f"Audit CSV: {path} ({len(entries)} zapisa)")
        return path

    def export_json(self, entries: List[Dict], output_path: str = "") -> str:
        """Export u JSON."""
        path = output_path or f"data/exports/audit_{date.today().isoformat()}.json"
        Path(path).parent.mkdir(parents=True, exist_ok=True)

        with open(path, "w", encoding="utf-8") as f:
            json.dump({
                "generated": datetime.now().isoformat(),
                "count": len(entries),
                "entries": entries
            }, f, ensure_ascii=False, indent=2, default=str)

        return path

    def summary(self, entries: List[Dict]) -> Dict[str, Any]:
        """Generiraj sažetak audit trail-a."""
        if not entries:
            return {"total": 0}

        users = {}
        actions = {}
        modules = {}
        risk_counts = {"high": 0, "medium": 0, "low": 0}

        for e in entries:
            u = e.get("user", "unknown")
            users[u] = users.get(u, 0) + 1
            a = e.get("action", "unknown")
            actions[a] = actions.get(a, 0) + 1
            m = e.get("module", "unknown")
            modules[m] = modules.get(m, 0) + 1
            r = e.get("risk_level", "low")
            risk_counts[r] = risk_counts.get(r, 0) + 1

        return {
            "total": len(entries),
            "period": {
                "from": entries[-1].get("timestamp", "") if entries else "",
                "to": entries[0].get("timestamp", "") if entries else "",
            },
            "by_user": dict(sorted(users.items(), key=lambda x: -x[1])[:10]),
            "by_action": dict(sorted(actions.items(), key=lambda x: -x[1])[:10]),
            "by_module": dict(sorted(modules.items(), key=lambda x: -x[1])[:10]),
            "risk": risk_counts,
        }
